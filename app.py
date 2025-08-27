import os
import logging
import uuid
import tempfile
from flask import Flask, render_template, request, send_file, jsonify, flash, redirect, url_for
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
import pdf2docx
from pathlib import Path
import PyPDF2
import tabula
import pandas as pd

# Configure logging
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")

# Configuration
UPLOAD_FOLDER = 'uploads'
CONVERTED_FOLDER = 'converted'
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB max file size
ALLOWED_EXTENSIONS = {'pdf'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['CONVERTED_FOLDER'] = CONVERTED_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Ensure upload and converted directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(CONVERTED_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_pdf(file_path):
    """Validate that the uploaded file is a proper PDF."""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            # Check if PDF has pages and is readable
            if len(pdf_reader.pages) == 0:
                return False, "PDF file appears to be empty"
            
            # Try to read first page to ensure it's not corrupted
            first_page = pdf_reader.pages[0]
            first_page.extract_text()  # This will raise an exception if corrupted
            
            return True, "Valid PDF file"
    except Exception as e:
        logging.error(f"PDF validation error: {str(e)}")
        return False, f"Invalid or corrupted PDF file: {str(e)}"

def convert_pdf_to_docx(pdf_path, docx_path):
    """Convert PDF to DOCX using pdf2docx library with enhanced error handling."""
    try:
        logging.info(f"Starting conversion: {pdf_path} -> {docx_path}")
        
        # Validate input file exists and is readable
        if not os.path.exists(pdf_path):
            return False, "Input PDF file not found"
        
        if os.path.getsize(pdf_path) == 0:
            return False, "PDF file is empty"
        
        # Use pdf2docx converter with specific settings for better compatibility
        cv = pdf2docx.Converter(pdf_path)
        
        # Convert with enhanced settings to preserve formatting
        cv.convert(
            docx_path, 
            start=0, 
            end=None,
            multi_processing=False,  # Disable multiprocessing for stability
            cpu_count=1  # Use single thread for more reliable conversion
        )
        cv.close()
        
        # Validate output file was created and has content
        if not os.path.exists(docx_path):
            return False, "DOCX file was not created"
        
        if os.path.getsize(docx_path) == 0:
            return False, "Generated DOCX file is empty"
        
        # Basic validation - try to read the file as a DOCX
        try:
            from zipfile import ZipFile, BadZipFile
            with ZipFile(docx_path, 'r') as zip_file:
                # Check if it contains the required DOCX structure
                required_files = ['[Content_Types].xml', 'word/document.xml']
                zip_contents = zip_file.namelist()
                
                for required_file in required_files:
                    if required_file not in zip_contents:
                        logging.warning(f"DOCX structure check: Missing {required_file}")
                        
        except BadZipFile:
            return False, "Generated file is not a valid DOCX format"
        except Exception as validation_error:
            logging.warning(f"DOCX validation warning: {str(validation_error)}")
            # Continue anyway as the file might still be usable
        
        logging.info(f"Conversion completed successfully: {docx_path}")
        return True, "Conversion completed successfully"
        
    except ImportError as e:
        logging.error(f"Missing required library: {str(e)}")
        return False, "Conversion library not available. Please contact support."
    except MemoryError:
        logging.error("Memory error during conversion")
        return False, "File too large or complex for conversion. Try a smaller file."
    except Exception as e:
        logging.error(f"Conversion error: {str(e)}")
        error_msg = str(e)
        
        # Provide more user-friendly error messages
        if "No module named" in error_msg:
            return False, "Required conversion libraries are not installed"
        elif "Permission denied" in error_msg:
            return False, "File access permission error"
        elif "No such file" in error_msg:
            return False, "Input file could not be found"
        else:
            return False, f"Conversion failed: {error_msg}"

def convert_pdf_to_excel(pdf_path, xlsx_path):
    """Convert PDF to Excel using tabula-py library with enhanced formatting preservation."""
    try:
        logging.info(f"Starting PDF to Excel conversion: {pdf_path} -> {xlsx_path}")
        
        # Validate input file exists and is readable
        if not os.path.exists(pdf_path):
            return False, "Input PDF file not found"
        
        if os.path.getsize(pdf_path) == 0:
            return False, "PDF file is empty"
        
        # Extract all tables from the PDF with enhanced header detection
        try:
            # First attempt: Preserve headers with explicit header detection
            dfs = tabula.read_pdf(
                pdf_path, 
                pages='all',
                multiple_tables=True,
                pandas_options={'header': [0]},  # Explicitly use first row as header
                lattice=True,
                stream=False,
                guess=False,
                area=None,
                relative_area=True,
                relative_columns=True
            )
            
            # Second attempt: Try with multiple header rows for complex tables
            if not dfs or len(dfs) == 0 or any(df.empty for df in dfs):
                dfs = tabula.read_pdf(
                    pdf_path, 
                    pages='all',
                    multiple_tables=True,
                    pandas_options={'header': [0, 1]},  # Try multi-row headers
                    lattice=True,
                    stream=False,
                    guess=False
                )
            
            # Third attempt: Stream mode with header preservation
            if not dfs or len(dfs) == 0 or any(df.empty for df in dfs):
                dfs = tabula.read_pdf(
                    pdf_path, 
                    pages='all',
                    multiple_tables=True,
                    pandas_options={'header': 0},  # Single header row
                    stream=True,
                    guess=True,
                    relative_area=True,
                    relative_columns=True
                )
            
            # Fourth attempt: Raw extraction with manual header handling
            if not dfs or len(dfs) == 0 or any(df.empty for df in dfs):
                dfs = tabula.read_pdf(
                    pdf_path,
                    pages='all',
                    multiple_tables=True,
                    pandas_options={'header': None},  # No automatic header
                    lattice=False,
                    stream=True,
                    guess=True
                )
                
        except Exception as tabula_error:
            logging.warning(f"Tabula extraction error: {str(tabula_error)}")
            # Last resort: try to extract any data
            dfs = []
            
        # Process and clean up the dataframes with enhanced header handling
        processed_dfs = []
        
        for i, df in enumerate(dfs):
            if df is not None and not df.empty:
                # Handle multi-level column headers
                if hasattr(df.columns, 'nlevels') and df.columns.nlevels > 1:
                    # Flatten multi-level columns
                    df.columns = [
                        ' '.join([str(c) for c in col if not pd.isna(c) and str(c).strip() != '']) 
                        if isinstance(col, tuple) else str(col)
                        for col in df.columns
                    ]
                
                # Clean up column names - preserve original names when possible
                new_columns = []
                first_row_used_as_header = False
                
                for j, col in enumerate(df.columns):
                    col_str = str(col).strip()
                    
                    # Check if this looks like an actual header (not just "Unnamed")
                    if (col_str.startswith('Unnamed') or 
                        pd.isna(col) or 
                        col_str == '' or
                        col_str == 'nan' or
                        col_str.lower() == 'none'):
                        
                        # Try to use the first row as header if columns are unnamed
                        if len(df) > 0 and j < len(df.columns):
                            potential_header = str(df.iloc[0, j]).strip()
                            if (potential_header and 
                                potential_header != 'nan' and 
                                potential_header.lower() != 'none' and 
                                len(potential_header) > 0 and
                                not potential_header.isdigit()):  # Avoid using numbers as headers
                                new_columns.append(potential_header)
                                first_row_used_as_header = True
                            else:
                                new_columns.append(f"Column_{j+1}")
                        else:
                            new_columns.append(f"Column_{j+1}")
                    else:
                        new_columns.append(col_str)
                
                df.columns = new_columns
                
                # Remove the first row if it was used for headers
                if first_row_used_as_header and len(df) > 0:
                    df = df.iloc[1:].reset_index(drop=True)
                
                # Remove completely empty rows and columns
                df = df.dropna(axis=0, how='all')  # Remove empty rows
                df = df.dropna(axis=1, how='all')  # Remove empty columns
                
                # Only keep dataframes that have actual data
                if len(df) > 0 and len(df.columns) > 0:
                    processed_dfs.append(df)
        
        # Update the dataframes list
        dfs = processed_dfs
        
        # Check if we got any data after processing
        if not dfs or len(dfs) == 0:
            return False, "No tables or data found in the PDF. The file might contain only images or be a scanned document."
        
        # Create Excel writer with formatting options
        with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
            for i, df in enumerate(dfs):
                # Clean up the dataframe
                if df is not None and not df.empty:
                    # Remove completely empty rows and columns
                    df = df.dropna(how='all').dropna(axis=1, how='all')
                    
                    # If dataframe still has data, write it to Excel
                    if not df.empty:
                        sheet_name = f'Table_{i+1}' if len(dfs) > 1 else 'Data'
                        
                        # Ensure sheet name is valid (max 31 chars, no special chars)
                        sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_')
                        
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Get the worksheet to apply formatting
                        worksheet = writer.sheets[sheet_name]
                        
                        # Enhanced formatting and styling
                        from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
                        from openpyxl.utils import get_column_letter
                        
                        # Auto-adjust column widths with better calculation
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            for cell in column:
                                try:
                                    # Consider both content length and cell formatting
                                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                                    if cell_length > max_length:
                                        max_length = cell_length
                                except:
                                    pass
                            
                            # Set column width with better scaling
                            if max_length > 0:
                                adjusted_width = min(max(max_length + 4, 10), 60)  # Min 10, Max 60
                                worksheet.column_dimensions[column_letter].width = adjusted_width
                            else:
                                worksheet.column_dimensions[column_letter].width = 12
                        
                        # Define enhanced styling
                        header_font = Font(bold=True, size=11, name='Arial')
                        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        data_font = Font(size=10, name='Arial')
                        data_alignment = Alignment(vertical='center', wrap_text=False)
                        
                        # Enhanced border styles
                        thick_border = Border(
                            left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium')
                        )
                        
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        # Apply enhanced formatting to all cells
                        for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                                                        min_col=1, max_col=worksheet.max_column), 1):
                            for col_idx, cell in enumerate(row, 1):
                                if cell.value is not None:
                                    # Apply borders
                                    cell.border = thick_border if row_idx == 1 else thin_border
                                    
                                    # Apply fonts and alignment
                                    if row_idx == 1:  # Header row
                                        cell.font = header_font
                                        cell.fill = header_fill
                                        cell.alignment = header_alignment
                                    else:  # Data rows
                                        cell.font = data_font
                                        cell.alignment = data_alignment
                                        
                                        # Try to detect numeric values for right alignment
                                        try:
                                            float(str(cell.value).replace(',', '').replace('$', ''))
                                            cell.alignment = Alignment(horizontal='right', vertical='center')
                                        except (ValueError, AttributeError):
                                            # Keep left alignment for text
                                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # Set row heights for better appearance
                        for row in range(1, worksheet.max_row + 1):
                            if row == 1:  # Header row
                                worksheet.row_dimensions[row].height = 25
                            else:  # Data rows
                                worksheet.row_dimensions[row].height = 18
                        
                        # Freeze header row for better navigation
                        if worksheet.max_row > 1:
                            worksheet.freeze_panes = 'A2'
                        
                        logging.info(f"Added sheet '{sheet_name}' with {len(df)} rows")
            
            # If no valid data was written, create an empty sheet with a message
            if not writer.sheets:
                df_empty = pd.DataFrame({'Message': ['No tabular data found in the PDF file.']})
                df_empty.to_excel(writer, sheet_name='Notice', index=False)
        
        # Validate output file was created and has content
        if not os.path.exists(xlsx_path):
            return False, "Excel file was not created"
        
        if os.path.getsize(xlsx_path) == 0:
            return False, "Generated Excel file is empty"
        
        logging.info(f"PDF to Excel conversion completed successfully: {xlsx_path}")
        return True, "PDF to Excel conversion completed successfully"
        
    except ImportError as e:
        logging.error(f"Missing required library for Excel conversion: {str(e)}")
        return False, "Excel conversion library not available. Please contact support."
    except MemoryError:
        logging.error("Memory error during PDF to Excel conversion")
        return False, "File too large or complex for conversion. Try a smaller file."
    except Exception as e:
        logging.error(f"PDF to Excel conversion error: {str(e)}")
        error_msg = str(e)
        
        # Provide more user-friendly error messages
        if "Java" in error_msg:
            return False, "Java runtime required for PDF processing is not available"
        elif "No tables found" in error_msg:
            return False, "No tables detected in the PDF. Make sure the PDF contains tabular data."
        elif "Permission denied" in error_msg:
            return False, "File access permission error"
        elif "No such file" in error_msg:
            return False, "Input file could not be found"
        else:
            return False, f"PDF to Excel conversion failed: {error_msg}"

@app.route('/')
def home():
    """Render the home/landing page."""
    return render_template('home.html')

@app.route('/pdf-tools')
def pdf_tools():
    """Render the PDF tools page."""
    return render_template('pdf_tools.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and conversion."""
    try:
        # Check if file is in request
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file selected'})
        
        file = request.files['file']
        
        # Check if file is selected
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        # Check file extension
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Only PDF files are allowed'})
        
        # Generate unique filename
        unique_id = str(uuid.uuid4())
        original_filename = secure_filename(file.filename or "document.pdf")
        pdf_filename = f"{unique_id}_{original_filename}"
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
        
        # Save uploaded file
        file.save(pdf_path)
        logging.info(f"File saved: {pdf_path}")
        
        # Validate PDF file
        is_valid, validation_message = validate_pdf(pdf_path)
        if not is_valid:
            # Clean up uploaded file
            os.remove(pdf_path)
            return jsonify({'success': False, 'error': validation_message})
        
        # Generate output filename
        docx_filename = f"{unique_id}_{Path(original_filename).stem}.docx"
        docx_path = os.path.join(app.config['CONVERTED_FOLDER'], docx_filename)
        
        # Convert PDF to DOCX
        conversion_success, conversion_message = convert_pdf_to_docx(pdf_path, docx_path)
        
        # Clean up uploaded PDF file
        os.remove(pdf_path)
        
        if not conversion_success:
            return jsonify({'success': False, 'error': conversion_message})
        
        # Return success with download URL
        return jsonify({
            'success': True, 
            'message': 'File converted successfully',
            'download_url': url_for('download_file', filename=docx_filename),
            'filename': f"{Path(original_filename).stem}.docx"
        })
        
    except RequestEntityTooLarge:
        return jsonify({'success': False, 'error': 'File too large. Maximum size is 50MB.'})
    except Exception as e:
        logging.error(f"Upload error: {str(e)}")
        # Clean up files if they exist
        try:
            if 'pdf_path' in locals() and 'pdf_path' in vars():
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
        except:
            pass
        return jsonify({'success': False, 'error': f'An error occurred: {str(e)}'})

@app.route('/upload-excel', methods=['POST'])
def upload_file_excel():
    """Handle file upload and conversion to Excel."""
    try:
        # Check if file is in request
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file selected'})
        
        file = request.files['file']
        
        # Check if file is selected
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'})
        
        # Check file extension
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Only PDF files are allowed'})
        
        # Generate unique filename
        unique_id = str(uuid.uuid4())
        original_filename = secure_filename(file.filename or "document.pdf")
        pdf_filename = f"{unique_id}_{original_filename}"
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
        
        # Save uploaded file
        file.save(pdf_path)
        logging.info(f"File saved for Excel conversion: {pdf_path}")
        
        # Validate PDF file
        is_valid, validation_message = validate_pdf(pdf_path)
        if not is_valid:
            # Clean up uploaded file
            os.remove(pdf_path)
            return jsonify({'success': False, 'error': validation_message})
        
        # Generate output filename
        xlsx_filename = f"{unique_id}_{Path(original_filename).stem}.xlsx"
        xlsx_path = os.path.join(app.config['CONVERTED_FOLDER'], xlsx_filename)
        
        # Convert PDF to Excel
        conversion_success, conversion_message = convert_pdf_to_excel(pdf_path, xlsx_path)
        
        # Clean up uploaded PDF file
        os.remove(pdf_path)
        
        if not conversion_success:
            return jsonify({'success': False, 'error': conversion_message})
        
        # Return success with download URL
        return jsonify({
            'success': True, 
            'message': 'PDF to Excel conversion completed successfully',
            'download_url': url_for('download_file', filename=xlsx_filename),
            'filename': f"{Path(original_filename).stem}.xlsx"
        })
        
    except RequestEntityTooLarge:
        return jsonify({'success': False, 'error': 'File too large. Maximum size is 50MB.'})
    except Exception as e:
        logging.error(f"Excel upload error: {str(e)}")
        # Clean up files if they exist
        try:
            if 'pdf_path' in locals() and 'pdf_path' in vars():
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
        except:
            pass
        return jsonify({'success': False, 'error': f'An error occurred: {str(e)}'})

@app.route('/download/<filename>')
def download_file(filename):
    """Handle file download."""
    try:
        file_path = os.path.join(app.config['CONVERTED_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            flash('File not found or expired', 'error')
            return redirect(url_for('index'))
        
        # Clean up file after a delay to allow download to complete
        import threading
        def delayed_cleanup():
            import time
            time.sleep(5)  # Wait 5 seconds to ensure download completes
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logging.info(f"Cleaned up file: {file_path}")
            except Exception as e:
                logging.error(f"Error cleaning up file: {str(e)}")
        
        # Start cleanup in background thread
        cleanup_thread = threading.Thread(target=delayed_cleanup)
        cleanup_thread.daemon = True
        cleanup_thread.start()
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename.split('_', 1)[1] if '_' in filename else filename,
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        
    except Exception as e:
        logging.error(f"Download error: {str(e)}")
        flash('Error downloading file', 'error')
        return redirect(url_for('index'))

@app.errorhandler(413)
def too_large(e):
    """Handle file too large error."""
    return jsonify({'success': False, 'error': 'File too large. Maximum size is 50MB.'}), 413

@app.errorhandler(500)
def internal_error(e):
    """Handle internal server errors."""
    logging.error(f"Internal server error: {str(e)}")
    return jsonify({'success': False, 'error': 'Internal server error occurred'}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
